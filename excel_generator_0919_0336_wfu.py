# 代码生成时间: 2025-09-19 03:36:11
import gradio as gr
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment

"""
Excel表格自动生成器程序
"""

class ExcelGenerator:
    def __init__(self):
        # 初始化Excel工作簿
        self.wb = Workbook()
        # 默认创建一个工作表
        self.ws = self.wb.active
        self.ws.title = "Sheet1"

    def create_header(self, header_data):
        """
        创建Excel表头
        :param header_data: 表头数据列表
        :return: None
        """
        for col_num, header in enumerate(header_data, start=1):
            self.ws.cell(row=1, column=col_num).value = header
            self.ws.cell(row=1, column=col_num).alignment = Alignment(horizontal='center')

    def add_row(self, row_data):
        """
        添加一行数据到Excel表格
        :param row_data: 一行数据列表
        :return: None
        """
        if not row_data:
            raise ValueError("Row data cannot be empty")
        for col_num, value in enumerate(row_data, start=1):
            self.ws.cell(row=self.ws.max_row + 1, column=col_num).value = value

    def save_excel(self, file_name):
        """
        保存Excel文件
        :param file_name: 文件名
        :return: None
        """
        self.wb.save(file_name)

    def load_excel(self, file_name):
        """
        加载Excel文件
        :param file_name: 文件名
        :return: pandas DataFrame
        """
        try:
            df = pd.read_excel(file_name)
            return df
        except Exception as e:
            raise ValueError(f"Failed to load Excel file: {e}")

# 创建Excel生成器实例
excel_gen = ExcelGenerator()

# Gradio界面函数
def generate_excel(header, rows):
    """
    生成Excel表格
    :param header: 表头，字符串列表，每个元素用逗号分隔
    :param rows: 表格数据，二维字符串列表
    :return: 生成的Excel文件
    """
    # 解析表头
    header_data = [h.strip() for h in header.split(",")]
    
    # 创建表头
    excel_gen.create_header(header_data)
    
    # 添加数据行
    for row_data in rows:
        excel_gen.add_row([r.strip() for r in row_data.split(",")])
    
    # 生成Excel文件并返回
    output_file_name = "generated_excel.xlsx"
    excel_gen.save_excel(output_file_name)
    return output_file_name

# Gradio界面配置
iface = gr.Interface(
    fn=generate_excel,
    inputs=[
        gr.Textbox(label="Excel Header", placeholder="Enter headers separated by comma"),
        gr.Textbox(label="Excel Rows", placeholder="Enter rows separated by comma, new row per line")
    ],
    outputs=gr.Textbox(label="Download Link"),
    title="Excel表格自动生成器",
    description="输入表头和行数据，自动生成Excel表格"
)

# 运行Gradio界面
iface.launch()